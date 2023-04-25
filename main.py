import time 
import pywhatkit
import pyautogui
from pynput.keyboard import Key, Controller
from tkinter import *
from openpyxl import Workbook,load_workbook
#pyautogui.alert('Merhaba, pyautogui!')

keyboard = Controller()
array = []
user = {}
nameArray = []
telarray = []


def readExcel():
    wb = load_workbook("isimler.xlsx")
    ws = wb.active
    for row in ws.iter_rows(min_row=1, min_col=1, max_row=70, max_col=2):
        for cell in row:
            if cell.value != None:
                array.append(cell.value)
            

    for i in range(len(array)): #0 dan eleman sayısının bir eksiği kadar 4 mesela 
        if i%2 == 0:
            nameArray.append(array[i])
        else:
            telarray.append(array[i]) 
    
    for i in range(len(nameArray)): #i=0
        user.setdefault(nameArray[i],telarray[i])

def send_whatsapp_message(msg: str,y: str):
    try:
        pywhatkit.sendwhatmsg_instantly(
            
            phone_no=y, 
            message=msg,
            tab_close=True
        )
        time.sleep(14)
        pyautogui.click()
        time.sleep(2)
        keyboard.press(Key.enter)
        keyboard.release(Key.enter)
        print("Message sent!")
    except Exception as e:
        print(str(e))


def tkinterIF():
    tk = Tk()
    tk.title("Whatsapp Bot Programı --Melybgcvn")
    tk.geometry("400x250")
    tk.resizable(False,False) 
    lbl = Label(tk, text="Hoşgeldin, lütfen yollamak istediğin mesajı başında Merhaba olmadan gir:",font="Helvetica 12", bg="purple", fg="yellow")
    lbl.place(x=0,y=40)
    lbl2 = Label(tk, text="Excel dosyanın isminin isimler.xlsx olduğundan emin ol !",font="Helvetica 12", bg="yellow", fg="purple")
    lbl2.place(x=0,y=70)
    
    
    def signUp():
        global answer
        answer = entry.get()
        for x, y in user.items():
            print(x,y)
            message = f"Merhaba {x}," + answer
            print(message)
            send_whatsapp_message(msg=message,y=y)
            
    entry = Entry(tk, width=40)
    entry.place(x=0,y=120)
    btn2 = Button(tk,
            text = "Çalıştır", 
            font="Times 12 bold",
            padx="25", pady="10", 
            bg="red", fg="blue", cursor="hand2",
            activeforeground="green", activebackground="black",
            command=signUp)
    btn2.place(x=125,y=200)
    tk.mainloop()
    
if __name__ == "__main__":
    readExcel()
    tkinterIF()

    
        

